import os
import jsonpath
import openpyxl
import requests

exl = 'test_case_api.xlsx'


def excel_master_data(excel_name, sheet_name):
    if os.path.exists(excel_name):
        test_excel = openpyxl.load_workbook(excel_name)
        sheet = test_excel[sheet_name]
        list1 = []
        max_row = sheet.max_row
        for i in range(2, max_row + 1, 1):
            dict1 = dict(
                case_id=sheet.cell(row=i, column=1).value,
                interface=sheet.cell(row=i, column=2).value,
                method=sheet.cell(row=i, column=4).value,
                url=sheet.cell(row=i, column=5).value,
                data=sheet.cell(row=i, column=6).value,
                expected=sheet.cell(row=i, column=7).value
            )
            list1.append(dict1)
        return list1
    else:
        return '文件不存在'


def update_excel_expected(exl_name, exl_sheet_name, value, value2, update_name):
    test_excle = openpyxl.load_workbook(exl_name)
    sheet = test_excle[exl_sheet_name]
    result = sheet.cell(row=value, column=value2)
    result.value = update_name
    test_excle.save(exl_name)
    return update_name


def public_post_no_token_head(method, body, url, head={"X-Lemonban-Media-Type": "lemonban.v2",
                                                       "Content-Type": "application/json"}):
    if method.lower() == 'post':
        res_register = requests.post(url=url, json=body, headers=head).json()
    if method.lower() == 'patch':
        res_register = requests.patch(url=url, json=body, headers=head).json()
    if method.lower() == 'get':
        res_register = requests.get(url=url, json=body, headers=head).json()
    return res_register


def automatic(excel, sheet, log=None):
    res_register = excel_master_data(excel, sheet)
    if not isinstance(res_register, str):
        list_1 = []
        for register in res_register:
            register_expected = register['expected']
            if register['interface'] == 'register' or register['interface'] == 'login':
                res = public_post_no_token_head(register['method'], eval(register['data']), register['url'])
            else:
                login_res = public_post_no_token_head(method='post',
                                                      url='http://api.lemonban.com/futureloan/member/login',
                                                      body=log)
                dl_id = jsonpath.jsonpath(login_res, '$..id')[0]
                token = jsonpath.jsonpath(login_res, '$..token')[0]
                token_head = {"X-Lemonban-Media-Type": "lemonban.v2",
                              "Content-Type": "application/json",
                              "Authorization": "Bearer" + " " + token
                              }
                if register['interface'] == 'loan_add':
                    cz_data = eval(register['data'])
                    cz_data['member_id'] = dl_id
                    res = public_post_no_token_head(register['method'], cz_data, register['url'], head=token_head)
                else:
                    res = public_post_no_token_head(register['method'], eval(register['data']), register['url'],
                                                    head=token_head)
            print('-' * 30)
            print('case_id: {}'.format(register['case_id']))
            print('预期结果{}'.format(eval(register_expected)))
            res_expected = {'code': res['code'], 'msg': res['msg']}
            print('实际结果{}'.format(res_expected))
            if eval(register_expected) == res_expected:
                print('通过')
                cg = update_excel_expected(excel, sheet, register['case_id'] + 1, 8, '通过')
                consequence = {'sheet': sheet, 'case_id': register['case_id'], 'result': cg}
                list_1.append(consequence)
            else:
                print('不通过')
                sb = update_excel_expected(excel, sheet, register['case_id'] + 1, 8, '不通过')
                consequence = {'sheet': sheet, 'case_id': register['case_id'], 'result': sb}
                list_1.append(consequence)
        return list_1
    else:
        return '文件不存在'


# 除了注册登录，其他操作都需先登录。请设置登录手机号和密码
# 普通账号
login_body = {
    "mobile_phone": "15815541555",
    "pwd": "lemon123456"
}
# 管理员账号
login0_body = {
    "mobile_phone": "15815541666",
    "pwd": "lemon123456"
}
# 注册-----------------------------------------
# print(automatic(exl, 'register'))
# 登录-----------------------------------------
# print(automatic(exl, 'login'))
# 充值-----------------------------------------
print(automatic(exl, 'recharge', login_body))
# 加标-----------------------------------------
# print(automatic(exl, 'loan_add', login_body))
# 审核-----------------------------------------
# print(automatic(exl, 'loan_audit', login0_body))
